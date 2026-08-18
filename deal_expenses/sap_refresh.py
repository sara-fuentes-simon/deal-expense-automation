"""SAP workbook refresh used by the Streamlit application."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MASTER_SHEET_NAMES = ("SAP Report", "SAP Invoices Report")
BSNY_SHEET_NAME = "SAP"
ENTITY_TO_COMPANY_CODE = {"BSNY": "6652", "SANCAP": "1428"}
COMPANY_CODE_TO_ENTITY = {value: key for key, value in ENTITY_TO_COMPANY_CODE.items()}
POSTING_DATE_HEADER = "Posting Date"
DOCUMENT_ID_HEADER = "Document ID"
COST_CENTER_HEADER = "Cost Center"
GROSS_AMOUNT_HEADER = "Gross Amount in Local Currency"
VIM_STATUS_HEADER = "VIM Process Status Text"
POSTED_VIM_STATUS = "POSTED"
IN_SCOPE_SHEET_NAME = "In Scope CCs"
IN_SCOPE_NAME_HEADER = "CC Name"
IN_SCOPE_COST_CENTER_HEADER = "SAP LA CC"
XL_UP = -4162
XL_TO_LEFT = -4159
XL_PASTE_FORMATS = -4122
XL_CALCULATION_AUTOMATIC = -4105


@dataclass(frozen=True)
class SapSourceRow:
    entity: str
    document_id: str | None
    posting_date: date
    values: dict[str, Any]

    @property
    def key(self) -> tuple[str, str] | None:
        return (self.entity, self.document_id) if self.document_id else None


@dataclass(frozen=True)
class SapRefreshResult:
    bsny_rows: int
    sancap_rows: int
    appended_rows: int


def _normalize(value: Any) -> str:
    return " ".join(str(value or "").strip().split()).upper()


def _normalize_document_id(value: Any) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        number = Decimal(text)
    except InvalidOperation:
        return text.upper()
    return str(number.quantize(Decimal("1"))) if number == number.to_integral_value() else format(number.normalize(), "f")


def _coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except (OverflowError, ValueError):
            return None
    for pattern in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(str(value).strip(), pattern).date()
        except ValueError:
            continue
    return None


def _decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except InvalidOperation as error:
        raise ValueError(f"Gross amount {value!r} is not numeric.") from error


def _read_headers(worksheet: Any) -> dict[str, int]:
    last_column = worksheet.Cells(1, worksheet.Columns.Count).End(XL_TO_LEFT).Column
    values = worksheet.Range(worksheet.Cells(1, 1), worksheet.Cells(1, last_column)).Value2
    values = values[0] if isinstance(values, tuple) else (values,)
    headers: dict[str, int] = {}
    duplicates: set[str] = set()
    for column, value in enumerate(values, start=1):
        header = _normalize(value)
        if not header:
            continue
        if header in headers:
            duplicates.add(header)
        headers[header] = column
    if duplicates:
        raise ValueError(f"Duplicate headers in {worksheet.Name}: {sorted(duplicates)}")
    return headers


def _required_column(headers: dict[str, int], header: str, sheet_name: str) -> int:
    column = headers.get(_normalize(header))
    if column is None:
        raise ValueError(f"Required header {header!r} was not found in worksheet {sheet_name!r}.")
    return column


def _last_used_row(worksheet: Any, column: int = 1) -> int:
    return worksheet.Cells(worksheet.Rows.Count, column).End(XL_UP).Row


def _resolve_master_sheet(workbook: Any) -> Any:
    sheets = {_normalize(workbook.Worksheets(index).Name): workbook.Worksheets(index) for index in range(1, workbook.Worksheets.Count + 1)}
    for name in MASTER_SHEET_NAMES:
        worksheet = sheets.get(_normalize(name))
        if worksheet is not None:
            return worksheet
    raise ValueError(f"Master workbook must contain one of: {', '.join(MASTER_SHEET_NAMES)}.")


def _load_in_scope_cost_centers(workbook: Any) -> set[str]:
    worksheet = workbook.Worksheets(IN_SCOPE_SHEET_NAME)
    used_range = worksheet.UsedRange
    first_row, first_column = used_range.Row, used_range.Column
    last_row = first_row + used_range.Rows.Count - 1
    last_column = first_column + used_range.Columns.Count - 1
    values = worksheet.Range(worksheet.Cells(first_row, first_column), worksheet.Cells(last_row, last_column)).Value2
    if not isinstance(values, tuple):
        values = ((values,),)
    elif values and not isinstance(values[0], tuple):
        values = (values,)

    name_index = cost_center_index = None
    header_index = None
    for index, row in enumerate(values):
        headers = [_normalize(value) for value in row]
        if _normalize(IN_SCOPE_NAME_HEADER) in headers and _normalize(IN_SCOPE_COST_CENTER_HEADER) in headers:
            header_index = index
            name_index = headers.index(_normalize(IN_SCOPE_NAME_HEADER))
            cost_center_index = headers.index(_normalize(IN_SCOPE_COST_CENTER_HEADER))
            break
    if header_index is None or name_index is None or cost_center_index is None:
        raise ValueError(f"Could not find the in-scope table on {IN_SCOPE_SHEET_NAME!r}.")

    cost_centers = set()
    for row in values[header_index + 1:]:
        if all(value in (None, "") for value in row):
            break
        value = _normalize_document_id(row[cost_center_index])
        if value:
            cost_centers.add(value)
    if not cost_centers:
        raise ValueError(f"No SAP LA CC values were found on {IN_SCOPE_SHEET_NAME!r}.")
    return cost_centers


def _source_rows(worksheet: Any, entity: str, reporting_year: int) -> list[SapSourceRow]:
    headers = _read_headers(worksheet)
    posting_column = _required_column(headers, POSTING_DATE_HEADER, worksheet.Name)
    document_column = _required_column(headers, DOCUMENT_ID_HEADER, worksheet.Name)
    status_column = _required_column(headers, VIM_STATUS_HEADER, worksheet.Name)
    final_row = _last_used_row(worksheet)
    if final_row < 2:
        return []
    values = worksheet.Range(worksheet.Cells(2, 1), worksheet.Cells(final_row, max(headers.values()))).Value2
    if not isinstance(values, tuple):
        values = (values,)
    if values and not isinstance(values[0], tuple):
        values = (values,)

    rows = []
    for value_row in values:
        posting_date = _coerce_date(value_row[posting_column - 1])
        if posting_date is None or posting_date.year != reporting_year:
            continue
        if _normalize(value_row[status_column - 1]) != POSTED_VIM_STATUS:
            continue
        row_values = {header: value_row[column - 1] for header, column in headers.items()}
        row_values[_normalize(POSTING_DATE_HEADER)] = datetime.combine(posting_date, time.min)
        rows.append(SapSourceRow(entity, _normalize_document_id(value_row[document_column - 1]), posting_date, row_values))
    return rows


def _master_rows(worksheet: Any) -> tuple[dict[str, int], list[tuple[tuple[str, str] | None, date | None, str | None, Decimal]]]:
    headers = _read_headers(worksheet)
    entity_column = _required_column(headers, "Company Code", worksheet.Name)
    document_column = _required_column(headers, DOCUMENT_ID_HEADER, worksheet.Name)
    posting_column = _required_column(headers, POSTING_DATE_HEADER, worksheet.Name)
    cost_center_column = _required_column(headers, COST_CENTER_HEADER, worksheet.Name)
    gross_amount_column = _required_column(headers, GROSS_AMOUNT_HEADER, worksheet.Name)
    final_row = _last_used_row(worksheet, document_column)
    if final_row < 2:
        return headers, []
    values = worksheet.Range(worksheet.Cells(2, 1), worksheet.Cells(final_row, max(headers.values()))).Value2
    if not isinstance(values, tuple):
        values = (values,)
    if values and not isinstance(values[0], tuple):
        values = (values,)

    rows = []
    for value_row in values:
        company_code = _normalize_document_id(value_row[entity_column - 1])
        document_id = _normalize_document_id(value_row[document_column - 1])
        entity = COMPANY_CODE_TO_ENTITY.get(company_code)
        if company_code and document_id and entity is None:
            raise ValueError(f"Unsupported Company Code {company_code!r} in {worksheet.Name!r}.")
        rows.append(((entity, document_id) if entity and document_id else None, _coerce_date(value_row[posting_column - 1]), _normalize_document_id(value_row[cost_center_column - 1]), _decimal(value_row[gross_amount_column - 1])))
    return headers, rows


def _validate_controls(source_rows: list[SapSourceRow], master_rows: list[tuple[tuple[str, str] | None, date | None, str | None, Decimal]], reporting_year: int, scoped_cost_centers: set[str]) -> list[SapSourceRow]:
    source_counts = Counter(row.key for row in source_rows if row.key)
    duplicates = [key for key, count in source_counts.items() if count > 1]
    missing_ids = [row for row in source_rows if row.key is None]
    master_keys = [key for key, _, _, _ in master_rows if key]
    duplicate_master = [key for key, count in Counter(master_keys).items() if count > 1]
    if duplicate_master or duplicates or missing_ids:
        raise ValueError("SAP controls failed: duplicate or missing Document ID values were found.")

    source_by_entity: dict[str, list[SapSourceRow]] = defaultdict(list)
    master_by_entity: dict[str, list[tuple[date | None, str | None, Decimal]]] = defaultdict(list)
    for row in source_rows:
        source_by_entity[row.entity].append(row)
    for key, posting_date, cost_center, gross_amount in master_rows:
        if key:
            master_by_entity[key[0]].append((posting_date, cost_center, gross_amount))
    for entity in set(source_by_entity) | set(master_by_entity):
        source_entity_rows = source_by_entity[entity]
        master_entity_rows = [row for row in master_by_entity[entity] if row[0] and row[0].year == reporting_year]
        source_total = sum((_decimal(row.values.get(_normalize(GROSS_AMOUNT_HEADER))) for row in source_entity_rows if _normalize_document_id(row.values.get(_normalize(COST_CENTER_HEADER))) in scoped_cost_centers), Decimal("0"))
        master_total = sum((gross_amount for _, cost_center, gross_amount in master_entity_rows if cost_center in scoped_cost_centers), Decimal("0"))
        if len(source_entity_rows) < len(master_entity_rows) or source_total < master_total:
            raise ValueError(f"SAP controls failed for {entity}: source year-to-date data is lower than the master.")

    master_key_set = set(master_keys)
    return [row for row in source_rows if row.key not in master_key_set]


def _resize_table(worksheet: Any, old_last_row: int, new_last_row: int) -> None:
    for index in range(1, worksheet.ListObjects.Count + 1):
        table = worksheet.ListObjects(index)
        if table.Range.Row + table.Range.Rows.Count - 1 == old_last_row:
            first_column = table.Range.Column
            last_column = first_column + table.Range.Columns.Count - 1
            table.Resize(worksheet.Range(worksheet.Cells(table.Range.Row, first_column), worksheet.Cells(new_last_row, last_column)))
            return


def _append_rows(worksheet: Any, headers: dict[str, int], rows: list[SapSourceRow]) -> None:
    if not rows:
        return
    document_column = _required_column(headers, DOCUMENT_ID_HEADER, worksheet.Name)
    old_last_row = _last_used_row(worksheet, document_column)
    first_target_row = max(2, old_last_row + 1)
    last_target_row = first_target_row + len(rows) - 1
    final_column = max(headers.values())
    worksheet.Range(worksheet.Cells(old_last_row, 1), worksheet.Cells(old_last_row, final_column)).Copy()
    worksheet.Range(worksheet.Cells(first_target_row, 1), worksheet.Cells(last_target_row, final_column)).PasteSpecial(Paste=XL_PASTE_FORMATS)
    worksheet.Application.CutCopyMode = False

    company_column = _required_column(headers, "Company Code", worksheet.Name)
    worksheet.Range(worksheet.Cells(first_target_row, company_column), worksheet.Cells(last_target_row, company_column)).Value2 = tuple((ENTITY_TO_COMPANY_CODE[row.entity],) for row in rows)
    for header, column in headers.items():
        if header in {_normalize("Company Code"), _normalize("Year")} or not any(header in row.values for row in rows):
            continue
        worksheet.Range(worksheet.Cells(first_target_row, column), worksheet.Cells(last_target_row, column)).Value2 = tuple((row.values.get(header),) for row in rows)
    year_column = _required_column(headers, "Year", worksheet.Name)
    worksheet.Range(worksheet.Cells(first_target_row, year_column), worksheet.Cells(last_target_row, year_column)).Value2 = tuple((row.posting_date.year,) for row in rows)
    for column in range(1, final_column + 1):
        formula = worksheet.Cells(old_last_row, column).FormulaR1C1
        if isinstance(formula, str) and formula.startswith("="):
            worksheet.Range(worksheet.Cells(first_target_row, column), worksheet.Cells(last_target_row, column)).FormulaR1C1 = formula
    _resize_table(worksheet, old_last_row, last_target_row)


def _validate_output(output_path: Path) -> None:
    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        worksheet = next((workbook[name] for name in MASTER_SHEET_NAMES if name in workbook.sheetnames), None)
        if worksheet is None:
            raise ValueError("Saved output is missing its SAP worksheet.")
        headers = {_normalize(cell.value): index for index, cell in enumerate(worksheet[1], start=1) if _normalize(cell.value)}
        company_column = _required_column(headers, "Company Code", worksheet.title)
        document_column = _required_column(headers, DOCUMENT_ID_HEADER, worksheet.title)
        keys = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            entity = COMPANY_CODE_TO_ENTITY.get(_normalize_document_id(row[company_column - 1]))
            document_id = _normalize_document_id(row[document_column - 1])
            if entity and document_id:
                keys.append((entity, document_id))
        if any(count > 1 for count in Counter(keys).values()):
            raise ValueError("Saved SAP output contains duplicate Company Code + Document ID values.")
    finally:
        workbook.close()


def refresh_sap_workbook(master_path: Path, bsny_path: Path, sancap_path: Path, reporting_year: int) -> SapRefreshResult:
    """Append validated SAP rows to ``master_path`` in place using Microsoft Excel."""
    try:
        import pythoncom
        import win32com.client as win32
    except ImportError as error:
        raise RuntimeError("pywin32 is required for the SAP refresh.") from error

    excel = master_book = bsny_book = sancap_book = None
    refresh_succeeded = False
    pythoncom.CoInitialize()
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        master_book = excel.Workbooks.Open(str(master_path.resolve()), ReadOnly=False)
        bsny_book = excel.Workbooks.Open(str(bsny_path.resolve()), ReadOnly=True)
        sancap_book = excel.Workbooks.Open(str(sancap_path.resolve()), ReadOnly=True)
        if sancap_book.Worksheets.Count != 1:
            raise ValueError("The SANCAP SAP report must contain exactly one worksheet.")
        master_sheet = _resolve_master_sheet(master_book)
        bsny_sheet = bsny_book.Worksheets(BSNY_SHEET_NAME)
        sancap_sheet = sancap_book.Worksheets(1)
        scoped_cost_centers = _load_in_scope_cost_centers(master_book)
        headers, master_rows = _master_rows(master_sheet)
        bsny_rows = _source_rows(bsny_sheet, "BSNY", reporting_year)
        sancap_rows = _source_rows(sancap_sheet, "SANCAP", reporting_year)
        if not bsny_rows or not sancap_rows:
            raise ValueError(f"No posted {reporting_year} SAP data was found in one or both source workbooks.")
        new_rows = _validate_controls([*bsny_rows, *sancap_rows], master_rows, reporting_year, scoped_cost_centers)
        _append_rows(master_sheet, headers, new_rows)
        excel.Calculation = XL_CALCULATION_AUTOMATIC
        excel.CalculateFull()
        master_book.Save()
        refresh_succeeded = True
        return SapRefreshResult(len(bsny_rows), len(sancap_rows), len(new_rows))
    finally:
        for workbook in (master_book, bsny_book, sancap_book):
            if workbook is not None:
                workbook.Close(SaveChanges=refresh_succeeded and workbook is master_book)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()
    
