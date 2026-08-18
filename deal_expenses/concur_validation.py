"""Structural validation for the Concur section of the master workbook."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from deal_expenses.models import ValidationResult
from deal_expenses.sources.base_concur import EXPENSE_HEADER, SOURCE_TO_MASTER, YEAR_HEADER, normalize_header


class ConcurMasterWorkbookValidator:
    """Verify the master workbook can safely receive Concur source rows."""

    sheet_name = "Concur Report"
    helper_headers = (
        "First & Last Name",
        "CC Expense is Mapped to",
        "LOB",
        "In Scope",
        "Error",
    )

    def validate(self, master_path: Path) -> list[ValidationResult]:
        try:
            workbook = load_workbook(master_path, read_only=True, data_only=False)
            try:
                if self.sheet_name not in workbook.sheetnames:
                    return [ValidationResult("Concur master workbook", False, f"Worksheet '{self.sheet_name}' was not found.")]
                worksheet = workbook[self.sheet_name]
                header_columns = defaultdict(list)
                for column_number, cell in enumerate(worksheet[1], start=1):
                    header = normalize_header(cell.value)
                    if header:
                        header_columns[header].append(column_number)

                required_headers = (*self.helper_headers, *SOURCE_TO_MASTER.values(), YEAR_HEADER, EXPENSE_HEADER)
                invalid_headers = [header for header in required_headers if len(header_columns[header]) != 1]
                if invalid_headers:
                    details = []
                    for header in invalid_headers:
                        columns = header_columns[header]
                        details.append(f"'{header}' is missing" if not columns else f"'{header}' is repeated")
                    return [ValidationResult("Concur master workbook", False, "; ".join(details) + ".")]

                missing_formulas = []
                for helper_header in self.helper_headers:
                    formula = worksheet.cell(row=2, column=header_columns[helper_header][0]).value
                    if not isinstance(formula, str) or not formula.startswith("="):
                        missing_formulas.append(helper_header)
                if missing_formulas:
                    return [
                        ValidationResult(
                            "Concur master workbook",
                            False,
                            "Missing row 2 formula template for: " + ", ".join(missing_formulas),
                        )
                    ]
            finally:
                workbook.close()
        except Exception as error:
            return [ValidationResult("Concur master workbook", False, str(error))]

        return [ValidationResult("Concur master workbook", True, "Workbook structure and helper formulas are valid.")]
