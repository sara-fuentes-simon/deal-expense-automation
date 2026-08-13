"""Contracts and reusable preflight logic for source workbook formats."""

from __future__ import annotations

from abc import ABC, abstractmethod
from pathlib import Path

from openpyxl import load_workbook

from deal_expenses.models import SourceSummary, ValidationResult


YEAR_HEADER = "Year"
EXPENSE_HEADER = "Expense Amount (reimbursement currency)"
SOURCE_TO_MASTER = {
    "Custom 41 - Name": "Client Name",
    "Custom 42 - Name": "Project Name",
    "Custom 43 - Name": "Epense Name",
}


def normalize_header(value: object) -> str:
    return "" if value is None else str(value).strip()


def is_reporting_year(value: object, year: int) -> bool:
    try:
        return float(value) == float(year)
    except (TypeError, ValueError):
        return str(value).strip() == str(year)


class SourceAdapter(ABC):
    """Defines how one source workbook type is validated and summarized."""

    key: str
    display_name: str
    required_columns = (YEAR_HEADER, EXPENSE_HEADER, *SOURCE_TO_MASTER)

    @abstractmethod
    def select_worksheet(self, workbook):
        """Return the worksheet containing source expense rows."""

    def validate(self, source_path: Path) -> list[ValidationResult]:
        try:
            workbook = load_workbook(source_path, read_only=True, data_only=False)
            try:
                worksheet = self.select_worksheet(workbook)
                headers = [normalize_header(cell.value) for cell in worksheet[1]]
            finally:
                workbook.close()
        except Exception as error:
            return [ValidationResult(self.display_name, False, str(error))]

        missing_columns = [header for header in self.required_columns if header not in headers]
        results = []
        if missing_columns:
            results.append(
                ValidationResult(
                    self.display_name,
                    False,
                    "Missing required column(s): " + ", ".join(missing_columns),
                )
            )
        if not results:
            results.append(ValidationResult(self.display_name, True, "Workbook structure is valid."))
        return results

    def summarize(self, source_path: Path, year: int) -> SourceSummary:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            worksheet = self.select_worksheet(workbook)
            headers = [normalize_header(cell.value) for cell in worksheet[1]]
            year_index = headers.index(YEAR_HEADER)
            expense_index = headers.index(EXPENSE_HEADER)
            row_count = 0
            expense_total = 0.0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if is_reporting_year(row[year_index], year):
                    row_count += 1
                    expense_value = row[expense_index]
                    if expense_value not in (None, ""):
                        expense_total += float(expense_value)
        finally:
            workbook.close()

        if row_count == 0:
            raise ValueError(f"{self.display_name} has no {year} rows.")
        return SourceSummary(self.key, self.display_name, row_count, expense_total)
