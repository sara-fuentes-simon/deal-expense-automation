"""Contracts and reusable preflight logic for SAP source workbook formats."""

from __future__ import annotations

from abc import ABC, abstractmethod
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from deal_expenses.models import SourceSummary, ValidationResult


POSTING_DATE_HEADER = "Posting Date"
DOCUMENT_ID_HEADER = "Document ID"
GROSS_AMOUNT_HEADER = "Gross Amount in Local Currency"
COST_CENTER_HEADER = "Cost Center"
VIM_STATUS_HEADER = "VIM Process Status Text"
POSTED_VIM_STATUS = "POSTED"


def normalize_sap_header(value: object) -> str:
    """Return an uppercase, whitespace-normalized SAP header or value."""
    return " ".join(str(value or "").strip().split()).upper()


def is_sap_reporting_year(value: object, year: int) -> bool:
    """Return whether an SAP posting date belongs to the reporting year."""
    if isinstance(value, datetime):
        return value.year == year
    if isinstance(value, date):
        return value.year == year
    return False


class SapSourceAdapter(ABC):
    """Defines how one SAP source workbook is validated and summarized."""

    key: str
    display_name: str
    required_columns = (
        POSTING_DATE_HEADER,
        DOCUMENT_ID_HEADER,
        GROSS_AMOUNT_HEADER,
        COST_CENTER_HEADER,
        VIM_STATUS_HEADER,
    )

    @abstractmethod
    def select_worksheet(self, workbook):
        """Return the worksheet containing SAP source rows."""

    def validate(self, source_path: Path) -> list[ValidationResult]:
        try:
            workbook = load_workbook(source_path, read_only=True, data_only=False)
            try:
                worksheet = self.select_worksheet(workbook)
                headers = {normalize_sap_header(cell.value) for cell in worksheet[1]}
            finally:
                workbook.close()
        except Exception as error:
            return [ValidationResult(self.display_name, False, str(error))]

        missing_columns = [header for header in self.required_columns if normalize_sap_header(header) not in headers]
        if missing_columns:
            return [
                ValidationResult(
                    self.display_name,
                    False,
                    "Missing required column(s): " + ", ".join(missing_columns),
                )
            ]
        return [ValidationResult(self.display_name, True, "Workbook structure is valid.")]

    def summarize(self, source_path: Path, year: int) -> SourceSummary:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            worksheet = self.select_worksheet(workbook)
            headers = [normalize_sap_header(cell.value) for cell in worksheet[1]]
            posting_date_index = headers.index(normalize_sap_header(POSTING_DATE_HEADER))
            gross_amount_index = headers.index(normalize_sap_header(GROSS_AMOUNT_HEADER))
            vim_status_index = headers.index(normalize_sap_header(VIM_STATUS_HEADER))
            row_count = 0
            gross_total = 0.0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not is_sap_reporting_year(row[posting_date_index], year):
                    continue
                if normalize_sap_header(row[vim_status_index]) != POSTED_VIM_STATUS:
                    continue
                row_count += 1
                if row[gross_amount_index] not in (None, ""):
                    gross_total += float(row[gross_amount_index])
        finally:
            workbook.close()

        if row_count == 0:
            raise ValueError(f"{self.display_name} has no posted {year} rows.")
        return SourceSummary(self.key, self.display_name, row_count, gross_total)
