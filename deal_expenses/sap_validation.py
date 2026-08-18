"""Structural validation for the SAP section of the master workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from deal_expenses.models import ValidationResult
from deal_expenses.sources.base_sap import (
    COST_CENTER_HEADER,
    DOCUMENT_ID_HEADER,
    GROSS_AMOUNT_HEADER,
    POSTING_DATE_HEADER,
    normalize_sap_header,
)


class SapMasterWorkbookValidator:
    """Verify the master workbook can safely receive SAP source rows."""

    sheet_names = ("SAP Report", "SAP Invoices Report")
    in_scope_sheet_name = "In Scope CCs"
    required_headers = (
        "Company Code",
        DOCUMENT_ID_HEADER,
        POSTING_DATE_HEADER,
        COST_CENTER_HEADER,
        GROSS_AMOUNT_HEADER,
        "Year",
    )

    def validate(self, master_path: Path) -> list[ValidationResult]:
        try:
            workbook = load_workbook(master_path, read_only=True, data_only=False)
            try:
                worksheet = next((workbook[name] for name in self.sheet_names if name in workbook.sheetnames), None)
                if worksheet is None:
                    return [
                        ValidationResult(
                            "SAP master workbook",
                            False,
                            f"Worksheet '{self.sheet_names[0]}' or '{self.sheet_names[1]}' was not found.",
                        )
                    ]
                if self.in_scope_sheet_name not in workbook.sheetnames:
                    return [
                        ValidationResult(
                            "SAP master workbook",
                            False,
                            f"Worksheet '{self.in_scope_sheet_name}' was not found.",
                        )
                    ]
                headers = [normalize_sap_header(cell.value) for cell in worksheet[1]]
            finally:
                workbook.close()
        except Exception as error:
            return [ValidationResult("SAP master workbook", False, str(error))]

        missing_headers = [header for header in self.required_headers if normalize_sap_header(header) not in headers]
        if missing_headers:
            return [
                ValidationResult(
                    "SAP master workbook",
                    False,
                    "Missing required column(s): " + ", ".join(missing_headers),
                )
            ]
        return [ValidationResult("SAP master workbook", True, "Workbook structure is valid.")]
